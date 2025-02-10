import streamlit as st
import random
import os
import base64
from openpyxl import load_workbook
from PIL import Image as PILImage
import io
import time  # For time tracking
import pymongo
from pymongo import MongoClient
from datetime import datetime

# Define categories
general_categories = [
    "aptitude",
    "data-interpretation",
    "verbal-ability",
    "logical-reasoning",
    "verbal-reasoning",
    "non-verbal-reasoning"
]

technical_categories = [
    "c-programming",
    "cpp-programming",
    "c-sharp-programming",
    "java-programming"
]

# Database connection setup (MongoDB)
def db_connect():
    client = MongoClient("mongodb://localhost:27017/")
    return client['quiz_system']

# Function to get the latest test number from the database
def get_test_number(username, category):
    db = db_connect()
    collection = db["apti_test"]
    latest_test_cursor = collection.find({"student_id": username, "category": category}).sort("timestamp", pymongo.DESCENDING).limit(1)
    latest_test = list(latest_test_cursor)
    if latest_test:
        test_no = latest_test[0]["test_no"] + 1
    else:
        test_no = 1
    return test_no

# Function to calculate individual test accuracy
def get_test_wise_accuracy(username, category, test_no):
    db = db_connect()
    collection = db['apti_test']
    test_details = collection.find({"student_id": username, "category": category, "test_no": test_no})
    correct_answers = 0
    total_questions = 0
    for test in test_details:
        correct_answers += test['marks_achieved']
        total_questions += test['no_of_questions']
    accuracy = (correct_answers / total_questions) * 100 if total_questions > 0 else 0
    return round(accuracy, 2)

# Function to calculate average accuracy
def get_average_accuracy(username, category, current_accuracy=None):
    db = db_connect()
    collection = db['apti_test']
    test_details = collection.find({"student_id": username, "category": category})
    total_accuracy = 0
    test_count = 0
    for test in test_details:
        test_accuracy = get_test_wise_accuracy(username, category, test['test_no'])
        total_accuracy += test_accuracy
        test_count += 1
    if current_accuracy is not None:
        total_accuracy += current_accuracy
    avg_test_accuracy = (total_accuracy / (test_count + 1)) if test_count > 0 else current_accuracy
    return round(avg_test_accuracy, 2)

# Store test details in the "apti_test" collection
def store_test_details(username, test_no, category, no_of_questions, marks_achieved, time_taken, avg_test_accuracy):
    db = db_connect()
    collection = db['apti_test']
    existing_test = collection.find_one({
        "student_id": username,
        "test_no": test_no,
        "category": category
    })
    if existing_test:
        # Do nothing if test details already exist.
        return
    test_data = {
        "student_id": username,
        "timestamp": datetime.now(),
        "category": category,
        "test_no": test_no,
        "no_of_questions": no_of_questions,
        "marks_achieved": marks_achieved,
        "time_taken": time_taken,
        "avg_test_accuracy": avg_test_accuracy
    }
    try:
        collection.insert_one(test_data)
        st.success("Test details stored successfully.")
    except Exception as e:
        st.error(f"Error inserting test data: {e}")

# Load questions from Excel files
def load_questions(category):
    questions = []
    category_list = general_categories if category == 'General' else technical_categories
    for subcategory in category_list:
        file_name = f"{subcategory}.xlsx"
        file_path = os.path.join(os.path.dirname(__file__), file_name)
        if not os.path.exists(file_path):
            continue  # Skip if the file doesn't exist
        wb = load_workbook(file_path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            question_no = row[0].value
            question_text = row[1].value
            options = row[2].value
            answer = row[3].value
            explanation = row[4].value
            img_path = None
            if question_text:
                for img in sheet._images:
                    if img.anchor._from.row == row[0].row - 1:
                        img_stream = io.BytesIO()
                        pil_image = PILImage.open(io.BytesIO(img._data()))
                        pil_image.save(img_stream, format='PNG')
                        img_stream.seek(0)
                        image_data = base64.b64encode(img_stream.read()).decode('utf-8')
                        img_path = f"data:image/png;base64,{image_data}"
                        break
            if question_no and question_text and options and answer:
                if subcategory == 'non-verbal-reasoning':
                    options_list = options.splitlines()
                else:
                    options_list = options.split(';')
                options_list = [option.strip() for option in options_list]
                labeled_options = {chr(65 + i): option for i, option in enumerate(options_list)}
                correct_label = None
                for label, option in labeled_options.items():
                    if option.strip().lower() == answer.strip().lower():
                        correct_label = label
                        break
                if correct_label is None:
                    correct_label = "Unknown"
                questions.append({
                    'question_no': question_no,
                    'question_text': question_text,
                    'image_data': img_path,
                    'options': options_list,
                    'labeled_options': labeled_options,
                    'correct_answer': correct_label,
                    'explanation': explanation.strip() if explanation else "No explanation available."
                })
    return questions

# Helper function to rerun the Streamlit app
def rerun_app():
    if hasattr(st, 'rerun'):
        st.rerun()
    elif hasattr(st, 'experimental_rerun'):
        st.experimental_rerun()
    else:
        st.error("Rerun not supported in this version of Streamlit. Please upgrade Streamlit.")

# --- Streamlit UI ---
st.title("🧠 AptiQuiz - Practice Your Skills!")

# Container for the initial form
with st.container():
    if "started" not in st.session_state or not st.session_state.started:
        username = st.text_input("Enter your username:")
        category = st.radio("Choose a category:", ['General', 'Technical'])
        if username:
            test_no = get_test_number(username, category)
            st.write(f"Test Number: {test_no}")
        test_type = st.radio("Select Your Test Mode:", ["⚡ Quick Challenge (10 Questions)", "🏆 Full Test (30 Questions)"])
        no_of_questions = 10 if "Quick Challenge" in test_type else 30
        if st.button("Start Quiz"):
            st.session_state.started = True
            st.session_state.username = username
            st.session_state.category = category
            st.session_state.test_no = test_no
            st.session_state.no_of_questions = no_of_questions
            st.session_state.questions = load_questions(category)
            random.shuffle(st.session_state.questions)
            st.session_state.questions = st.session_state.questions[:no_of_questions]
            st.session_state.current_question = 0
            st.session_state.score = 0
            st.session_state.user_answers = []
            st.session_state.start_time = time.time()  # Capture start time

# Divider before showing quiz questions
st.markdown("---")

# Display the quiz questions in their container
if "questions" in st.session_state and st.session_state.current_question < len(st.session_state.questions):
    with st.container():
        question_data = st.session_state.questions[st.session_state.current_question]
        with st.form(key=f"question_form_{st.session_state.current_question}"):
            st.header(f"Question {st.session_state.current_question + 1} / {len(st.session_state.questions)}")
            st.write(question_data["question_text"])
            if question_data["image_data"]:
                st.image(question_data["image_data"], use_container_width=True)
            options = question_data["labeled_options"]
            user_choice = st.radio(
                "Choose your answer:",
                list(options.keys()),
                format_func=lambda x: f"{x}: {options[x]}"
            )
            submit_button = st.form_submit_button(label="Submit Answer")
            if submit_button:
                st.session_state.user_answers.append(user_choice)
                if user_choice == question_data["correct_answer"]:
                    st.session_state.score += 1
                st.session_state.current_question += 1
                rerun_app()  # Force immediate rerun to update UI

# Show results in a separate container after the last question
if "questions" in st.session_state and st.session_state.current_question == len(st.session_state.questions):
    with st.container():
        st.markdown("---")
        end_time = time.time()
        time_taken = round(end_time - st.session_state.start_time, 2)
        st.header("🎉 Quiz Completed!")
        st.write(f"**Your Score:** {st.session_state.score} / {len(st.session_state.questions)}")
        st.write(f"**Time Taken:** {time_taken} seconds")
        st.subheader("Correct Answers and Explanations:")
        for i, q in enumerate(st.session_state.questions):
            user_ans = st.session_state.user_answers[i]
            correct_ans = q["correct_answer"]
            st.markdown(f"**Q{i + 1}:** {q['question_text']}")
            if q["image_data"]:
                st.image(q["image_data"], use_container_width=True)
            st.markdown(f"**✅ Correct Answer:** {correct_ans}")
            st.markdown(f"**❌ Your Answer:** {user_ans}")
            st.markdown(f"**💡 Explanation:** {q['explanation']}")
            st.write("---")
        current_accuracy = get_test_wise_accuracy(st.session_state.username, st.session_state.category, st.session_state.test_no)
        avg_test_accuracy = get_average_accuracy(st.session_state.username, st.session_state.category, current_accuracy)
        store_test_details(
            st.session_state.username,
            st.session_state.test_no,
            st.session_state.category,
            st.session_state.no_of_questions,
            st.session_state.score,
            time_taken,
            avg_test_accuracy
        )
        if st.button("Try Again"):
            st.session_state.started = False
            for key in ["username", "category", "test_no", "questions", "current_question", "score", "user_answers", "start_time"]:
                if key in st.session_state:
                    del st.session_state[key]
            rerun_app()
