from flask import Flask, render_template, request, redirect, url_for
import random
import json
import openpyxl
import base64
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import io

app = Flask(__name__)

# Define your categories here
general_categories = [
    #"aptitude", 
    #"data-interpretation", 
    #"verbal-ability", 
    "logical-reasoning", 
    "verbal-reasoning", 
    "non-verbal-reasoning"
]

technical_categories = [
    "c-programming", 
    "cpp-programming", 
    "c-sharp-programming", 
    "java-programming"
]

def load_questions(category):
    questions = []
    category_list = general_categories if category == 'general' else technical_categories
    
    for subcategory in category_list:
        file_path = f"{subcategory}.xlsx"  # Ensure the correct file path is set
        wb = load_workbook(file_path)
        sheet = wb.active
        
        for row in sheet.iter_rows(min_row=2):
            question_no = row[0].value  
            question_text = row[1].value  
            options = row[2].value  
            answer = row[3].value  
            explanation = row[4].value  

            # Initialize img_path to None
            img_path = None
            
            if question_text:
                for img in sheet._images:
                    if img.anchor._from.row == row[0].row - 1:  # Check if the image is in the same row
                        # Create a BytesIO stream to save the image
                        img_stream = io.BytesIO()
                        
                        # Convert the Excel image to a PIL image and save it to BytesIO
                        pil_image = PILImage.open(io.BytesIO(img._data()))
                        pil_image.save(img_stream, format='PNG')  # Use PIL to save the image
                        img_stream.seek(0)  # Reset stream position
                        image_data = base64.b64encode(img_stream.read()).decode('utf-8')
                        img_path = f"data:image/png;base64,{image_data}"  # Create the img path
                        break  # Exit the loop after finding the first image

            if question_no and question_text and options and answer: 
    
                if subcategory == 'non-verbal-reasoning':
                    # Split options by new lines, assuming they're vertically listed in Excel
                    options_list = options.splitlines()
                else:
                    options_list = options.split(';')

                # Clean up the options
                options_list = [option.strip() for option in options_list]

                # Use labeled options directly from the formatted options
                labeled_options = {chr(65 + i): option for i, option in enumerate(options_list)}

                # Find the correct label for the answer
                correct_label = None
                for label, option in labeled_options.items():
                    if option.strip().lower() == answer.strip().lower():
                        correct_label = label
                        break

                # Ensure 'correct_label' is always set
                if correct_label is None:
                    correct_label = "Unknown"

            if question_no and question_text and options and answer:  
                questions.append({
                    'question_no': question_no,
                    'question_text': question_text,
                    'image_data': img_path,  # Store the image path
                    'options': options.split(';'),  # Split options by semicolon
                    'labeled_options': labeled_options,  # Store labeled options
                    'answer': answer.strip(),  
                    'explanation': explanation.strip() if explanation else "No explanation available."
                })
    return questions


@app.route('/')
def index():
    return render_template('test.html')

@app.route('/quiz', methods=['POST'])
def quiz():
    category = request.form['category']
    selected_questions = load_questions(category)

    # Shuffle and select 20 questions
    random.shuffle(selected_questions)
    selected_questions = selected_questions[:5]

    # Start quiz with the first question
    return redirect(url_for('question', question_index=0, questions=json.dumps(selected_questions), score=0))

@app.route('/question/<int:question_index>', methods=['GET', 'POST'])
def question(question_index):
    questions_data = request.args.get('questions')
    score = int(request.args.get('score', 0))  # Get current score from query parameters

    # Deserialize questions data
    try:
        questions = json.loads(questions_data)
    except (json.JSONDecodeError, TypeError):
        return "Question data is missing!", 400

    # Initialize or retrieve user_answers
    user_answers_data = request.args.get('user_answers', '[]')  # Default to empty list if not found
    try:
        user_answers = json.loads(user_answers_data)
    except (json.JSONDecodeError, TypeError):
        user_answers = []

    # If POST method, handle the answer submission
    if request.method == 'POST':
        user_answer = request.form.get('user_answer')
        
        # Check if answer is provided
        if user_answer is None:
            user_answer = ' '

        # Store the answer for the current question
        user_answers.append(user_answer)  # Store as A, B, C, D

        # Calculate the score for the current question
        correct_answer = questions[question_index]['answer']
        
        # Check if the user's answer matches the correct answer
        if user_answer.strip().upper() == correct_answer.strip().upper():
            score += 1  # Increase score if the answer is correct
        
        # Prepare for the next question or results
        if question_index + 1 < len(questions):
            next_index = question_index + 1
            return redirect(url_for('question', question_index=next_index, questions=questions_data, score=score, user_answers=json.dumps(user_answers)))
        else:
            return redirect(url_for('results', score=score, total=len(questions), questions=json.dumps(questions), user_answers=json.dumps(user_answers)))

    # Display the current question
    if question_index < len(questions):
        current_question = questions[question_index]

        # Create a list of labeled options (A, B, C, D)
        labeled_options = {chr(65 + i): option for i, option in enumerate(current_question['options'])}
        
        return render_template('quiz.html', 
                               question=current_question, 
                               labeled_options=labeled_options,  # Pass labeled options
                               question_index=question_index + 1, 
                               total_questions=len(questions), 
                               questions=questions_data, 
                               score=score, 
                               user_answers=json.dumps(user_answers))  # Pass user answers
    else:
        return "Question index out of range!", 400

@app.route('/results', methods=['GET'])
def results():
    questions_data = request.args.get('questions')
    score = int(request.args.get('score', 0))
    user_answers_data = request.args.get('user_answers')

    if not questions_data or not user_answers_data:
        return "Question or user answers data is missing!", 400

    try:
        questions = json.loads(questions_data)
        user_answers = json.loads(user_answers_data)
    except json.JSONDecodeError:
        return "Error decoding questions or user answers data.", 400

    return render_template('results.html', score=score, total=len(questions), questions=questions, user_answers=user_answers, zip=zip)

if __name__ == '__main__':
    app.run(debug=True)
