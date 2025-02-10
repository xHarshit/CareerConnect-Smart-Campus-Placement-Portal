import requests
from bs4 import BeautifulSoup
import pandas as pd
import random
import os
from urllib.parse import urljoin
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def find_folders(folder_ls, url):
    try:
        # Fetch the content from the URL
        response = requests.get(url, verify=False)  # Set verify=False to ignore SSL warnings
        response.raise_for_status()  # Raise an exception for HTTP errors
        
        # Parse the content with BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Find all folder links
        folders = []
        for a_tag in soup.find_all('a', href=True):
            href = a_tag['href']
            if f'/{folder_ls}/' in href and href not in folders:
                folders.append(href)
        
        return folders

    except requests.RequestException as e:
        print(f"Error fetching content from {url}: {e}")
        return []

def handle_pages(base_url, total_pages, folder_name):
    # Create the new directory if it doesn't exist
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        #print(f"Created directory: {folder_name}")
    
    for page_number in range(1, total_pages + 1):
        # Adjust the page URL for the first page or subsequent pages
        if page_number == 1:
            page_url = base_url
        else:
            page_url = f"{base_url}/{str(page_number).zfill(6)}"
        
        #print(f"Processing page: {page_url}")
        
        # Retrieve content from the page
        data = retrieve_content(page_url)
        
        if data:
            # Construct the file name using the folder path and base URL
            file_name = os.path.join(folder_name, base_url.split('/')[-2] + '.xlsx')
            
            # Create the file or append to it
            create_or_append_xls_file(data, file_name)
        else:
            print(f"No content retrieved for {page_url}")

    print(f"All files saved in the directory: {folder_name}")

def create_or_append_xls_file(data, filename):
    try:
        if os.path.exists(filename):
            df_existing = pd.read_excel(filename)
            df_new = pd.DataFrame(data)
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            df_combined.to_excel(filename, index=False)
        else:
            df_new = pd.DataFrame(data)
            df_new.to_excel(filename, index=False)
        
        #print(f"Excel file '{filename}' updated successfully.")
    except Exception as e:
        print(f"Error creating or appending Excel file '{filename}': {e}")

def retrieve_content(url):
    try:
        response = requests.get(url, verify=False)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        
        questions = soup.find_all('div', class_='bix-div-container')
        data = []
        
        for question in questions:
            q_number = question.find('div', class_='bix-td-qno').text.strip()
            q_text = question.find('div', class_='bix-td-qtxt').text.strip()
            options_div = question.find('div', class_='bix-tbl-options')
            
            options = []
            if options_div:
                for option in options_div.find_all('div', class_='bix-opt-row'):
                    option_text = option.text.strip()
                    if option_text:
                        options.append(option_text)
            
            answer = question.find('input', class_='jq-hdnakq')['value']
            explanation = question.find('div', class_='bix-div-answer').text.strip() if question.find('div', class_='bix-div-answer') else 'No explanation provided'
            
            data.append({
                'Question Number': q_number,
                'Question Text': q_text,
                'Options': '; '.join(options),
                'Answer': answer,
                'Explanation': explanation
            })
        
        return data
    except requests.RequestException as e:
        print(f"Error fetching content from {url}: {e}")
        return []

def process_all_folders_and_subfolders(url, folder_name, total_pages=7):
    # Find and process all folders
    folders = find_folders(folder_name, url)

    for folder in folders:
        print(f"Processing folder: {folder}")
        
        # Find subfolders within the current folder
        subfolders = find_folders(folder_name, folder)
        
        # If no subfolders are found, handle just the main folder
        if not subfolders:
            handle_pages(folder, total_pages, folder_name)
        else:
            # Process each subfolder individually
            for subfolder in subfolders:
                print(f"Processing subfolder: {subfolder}")
                handle_pages(subfolder, total_pages, folder_name)

        # After processing all subfolders of the current folder, move to the next folder
        print(f"Finished processing folder: {folder}")
        break

    print("All folders and subfolders processed successfully.")

def combine_data(xlsx_files, output_file):
    all_data = []
    unique_questions = set()  # To keep track of unique questions

    for file_path in xlsx_files:
        folder_name = os.path.splitext(os.path.basename(file_path))[0]
        df = read_data_from_file(file_path)
        
        for index, row in df.iterrows():
            question = row.get('Question Text', '')
            if question and question not in unique_questions:
                unique_questions.add(question)
                all_data.append({
                    'Question Number': row.get('Question Number', ''),
                    'Question Text': question,
                    'Options': row.get('Options', ''),
                    'Answer': row.get('Answer', ''),
                    'Explanation': row.get('Explanation', ''),
                    'Folder Name': folder_name
                })
    
    # Create a DataFrame from the consolidated data
    consolidated_df = pd.DataFrame(all_data)

    # Write the DataFrame to an .xlsx file using openpyxl
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        consolidated_df.to_excel(writer, index=False, sheet_name='Consolidated Data')

def search_xlsx_files(directory):
    """Search for all .xlsx files in the given directory, excluding temporary files."""
    xlsx_files = [os.path.join(directory, f) for f in os.listdir(directory)
                  if f.endswith('.xlsx') and not f.startswith('~$')]
    return xlsx_files

def check_for_duplicates(file_path):
    """Check for duplicate entries in the Excel file based on 'Question Text'."""
    try:
        # Read the data from the Excel file
        df = pd.read_excel(file_path)

        # Check for duplicate questions
        if 'Question Text' not in df.columns:
            print("The expected column 'Question Text' is missing from the file.")
            return

        # Identify duplicates
        duplicates = df[df.duplicated(subset=['Question Text'], keep=False)]

        if duplicates.empty:
            print("No duplicate questions found.")
        else:
            print(f"Found {len(duplicates)} duplicate entries.")
            print(duplicates[['Question Number', 'Question Text']])
            # Optionally, save duplicates to a separate file
            duplicates.to_excel('duplicates_found.xlsx', index=False)
            print("Duplicate entries have been saved to 'duplicates_found.xlsx'.")

    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

def remove_duplicates_from_file(file_path):
    """Read an .xlsx file, remove duplicate rows, and save it back."""
    try:
        df = pd.read_excel(file_path)
        
        # Remove duplicate rows based on 'Question Number' and 'Question Text'
        df_cleaned = df.drop_duplicates(subset=['Question Number', 'Question Text'])
        
        # Save the cleaned DataFrame back to the file
        df_cleaned.to_excel(file_path, index=False)
        print(f"Duplicates removed and file saved: {file_path}")
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")

def find_last_non_empty_row(file_path):
    """Find the last non-empty row in an .xlsx file."""
    try:
        df = pd.read_excel(file_path)
        if not df.empty:
            # Find the last non-empty row index
            last_row_index = df.last_valid_index()
            return last_row_index
        else:
            return None
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return None

def process_all_xlsx_files_in_folder(folder_path):
    """Process all .xlsx files in the given folder and find the last non-empty row."""
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file_name)
            last_row_index = find_last_non_empty_row(file_path)
            if last_row_index is not None:
                print(f"{file_name} - No of Questions : {last_row_index-1}")
            else:
                print(f"File: {file_name} is empty or could not be processed.")


folder_list = ["aptitude", "data-interpretation", "verbal-ability", "logical-reasoning", "verbal-reasoning",
               "c-programming", "cpp-programming", "c-sharp-programming", "java-programming"]

for folder in folder_list:
    url = f'https://www.indiabix.com/{folder}/questions-and-answers/'
    folder_name = folder  # Name of the directory to save all files
    folder_path = "./{folder_name}"

    # Call the function to process folders and subfolders
    process_all_folders_and_subfolders(url, folder_name)

    directory = f'./{folder}/'  # Correctly format the directory path

    # Search for .xlsx files
    xlsx_files = search_xlsx_files(directory)

    # Output .xlsx file
    output_file = f'{folder}.xlsx'  # Correctly format the output file name

    # Consolidate data and write to the output file
    combine_data(xlsx_files, output_file)

    # Path to the current folder containing .xlsx files
current_folder = '.'  # Replace 'your_folder_name' with your folder path

# Process all .xlsx files in the current folder
process_all_xlsx_files_in_folder(current_folder)

# non-verbal -->

# Base URL for non-verbal reasoning section
base_url = "https://www.indiabix.com/non-verbal-reasoning/"
img_class = "nvr-image-opacity"

# Create a workbook and sheet for storing question data
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Questions"
# Add the new "Subtopic" column along with others
sheet.append(["Question Number", "Question Text", "Options", "Answer", "Explanation", "Folder Name"])


def extract_text(soup, class_name):
    """Helper function to extract text from a given class in BeautifulSoup."""
    element = soup.find(class_=class_name)
    return element.get_text(strip=True) if element else ""

def add_image_to_sheet(img_path, cell):
    """Add image to the Excel sheet at a specified cell."""
    if os.path.exists(img_path):
        img = XLImage(img_path)
        img.width = 100  # Adjust width
        img.height = 100  # Adjust height
        sheet.add_image(img, cell)
    else:
        print(f"Image not found: {img_path}")

def scrape_questions_from_page(url, folder_name, subtopic_name):
    data = retrieve_content(url)
    if data:
        soup = BeautifulSoup(data, 'html.parser')

        # Extract all question divs
        question_divs = soup.find_all('div', class_='bix-div-container')

        for question_div in question_divs:
            # Extract question number
            question_no = extract_text(question_div, 'bix-td-qno')

            # Extract question text and image
            question_text = extract_text(question_div, 'bix-td-qtxt')
            question_img = question_div.find('img', class_=img_class)
            question_img_src = question_img['src'] if question_img else None
            
            # Add question image to the Excel
            if question_img_src:
                question_img_path = os.path.join(folder_name, os.path.basename(question_img_src))
                question_text += f" [Image: {os.path.basename(question_img_src)}]"
                add_image_to_sheet(question_img_path, f"C{sheet.max_row + 1}")

            # Extract options (both text and images) with corresponding letters (A, B, C, D)
            options = []
            option_divs = question_div.find_all('div', class_='bix-td-option')
            option_letters = ['A', 'B', 'C', 'D', 'E']
            for idx, option_div in enumerate(option_divs):
                option_text = extract_text(option_div, 'bix-option')
                option_img = option_div.find('img', class_=img_class)
                option_img_src = option_img['src'] if option_img else None

                # Add option image to the Excel
                if option_img_src:
                    option_img_path = os.path.join(folder_name, os.path.basename(option_img_src))
                    option_text += f" [Image: {os.path.basename(option_img_src)}]"
                    add_image_to_sheet(option_img_path, f"D{sheet.max_row + 1}")
                
                options.append(f"{option_letters[idx]}. {option_text}")

            # Extract correct answer (text only)
            correct_answer = question_div.find('input', class_='jq-hdnakq')['value']

            # Extract explanation (if any) (text only)
            explanation = ""
            explanation_div = question_div.find('div', class_='bix-ans-description')
            if explanation_div:
                explanation = explanation_div.get_text(strip=True)

            # Append the data to the Excel sheet, including subtopic name
            sheet.append([question_no, question_text, "\n".join(options), correct_answer, explanation, subtopic_name])

def handle_pages_for_excel(base_url, total_pages, folder_name, subtopic_name):
    for page_number in range(1, total_pages + 1):
        page_url = base_url if page_number == 1 else f"{base_url}/{str(page_number).zfill(6)}"
        scrape_questions_from_page(page_url, folder_name, subtopic_name)

def scrape_images_from_folders(base_url, folder_name):
    folders = find_folders("non-verbal-reasoning", base_url)
    
    for folder in folders:
        folder_url = urljoin(base_url, folder)
        subfolders = find_folders("non-verbal-reasoning", folder_url)
        subtopic_name = folder.split('/')[-2]  # Extract subtopic name from URL
        
        if subfolders:
            for subfolder in subfolders:
                subfolder_url = urljoin(base_url, subfolder)
                subfolder_path = os.path.join(folder_name, subfolder.split('/')[-2])
                subtopic_name = subfolder.split('/')[-2]  # Update subtopic name for subfolder
                handle_pages_for_excel(subfolder_url, total_pages=10, folder_name=subfolder_path, subtopic_name=subtopic_name)
        else:
            folder_path = os.path.join(folder_name, folder.split('/')[-2])
            handle_pages_for_excel(folder_url, total_pages=10, folder_name=folder_path, subtopic_name=subtopic_name)

        print(f"Completed scraping folder: {folder}")
        
    print("All folders have been processed. Stopping.")


def remove_duplicates_and_clear_extra_rows(file_path):
    # Load the existing workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Create a set to track unique rows
    unique_rows = set()
    rows_to_keep = []
    last_filled_row_index = 0  # To track the last filled row index

    # Iterate through each row in the sheet
    for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
        # Convert row to a tuple (to be hashable)
        row_tuple = tuple(row)

        # Check if the row is unique
        if row_tuple not in unique_rows:
            unique_rows.add(row_tuple)  # Add to the set of unique rows
            rows_to_keep.append(row)     # Store the row for later writing
            last_filled_row_index = row_index + 1  # Update last filled row index

    # Clear all rows from the sheet
    sheet.delete_rows(1, sheet.max_row)

    # Write the unique rows back to the sheet
    for row in rows_to_keep:
        sheet.append(row)

    # Clear all rows below the last filled row
    if last_filled_row_index < sheet.max_row:
        sheet.delete_rows(last_filled_row_index + 1, sheet.max_row)

    # Save the modified workbook
    wb.save(file_path)

    print(f"Duplicates removed and content cleared from row {last_filled_row_index + 1} onwards. File saved.")

def process_all_xlsx_files(folder_path):
    """Process all .xlsx files in the specified folder."""
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file_name)
            remove_duplicates_from_file(file_path)

def read_data_from_file(file_path):
    """Read data from an .xlsx file and return it as a DataFrame."""
    try:
        df = pd.read_excel(file_path)
        return df
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return pd.DataFrame()  # Return an empty DataFrame
    except PermissionError:
        print(f"Permission denied: {file_path}")
        return pd.DataFrame()  # Return an empty DataFrame
    

base_folder = "./non-verbal-reasoning"

scrape_images_from_folders(base_url, base_folder)

# Save the Excel file
wb.save("non-verbal-0reasoning.xlsx")

file_path = "non-verbal-reasoning.xlsx"  # Your existing file path

remove_duplicates_and_clear_extra_rows(file_path)